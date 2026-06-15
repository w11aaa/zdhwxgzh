from __future__ import annotations

import re
import sqlite3
from pathlib import Path
from typing import Any

from PIL import Image, ImageDraw, ImageFont

from .config import CONFIG


OUTPUT_ROOT = CONFIG.project_root / "wechat_outputs" / "attachment_images"


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(CONFIG.database_path)
    conn.row_factory = sqlite3.Row
    return conn


def _font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        Path("C:/Windows/Fonts/msyh.ttc"),
        Path("C:/Windows/Fonts/simhei.ttf"),
        Path("C:/Windows/Fonts/simsun.ttc"),
    ]
    for path in candidates:
        if path.exists():
            return ImageFont.truetype(str(path), size=size)
    return ImageFont.load_default()


def _clean_line(value: str) -> str:
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", str(value or ""))
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont, max_width: int) -> list[str]:
    text = _clean_line(text)
    if not text:
        return [""]
    lines: list[str] = []
    current = ""
    for char in text:
        candidate = current + char
        if draw.textbbox((0, 0), candidate, font=font)[2] <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = char
    if current:
        lines.append(current)
    return lines or [""]


def _rows_from_parsed_text(text: str, max_rows: int) -> tuple[str, list[list[str]]]:
    lines = [_clean_line(line) for line in str(text or "").splitlines()]
    lines = [line for line in lines if line]
    if not lines:
        return "附件岗位表", []

    title = lines[0]
    header_idx = -1
    for i, line in enumerate(lines):
        if line in {"部门", "岗位"} and i + 3 < len(lines):
            window = " ".join(lines[i : i + 8])
            if "专业" in window and ("引进计划数" in window or "招聘人数" in window or "人数" in window):
                header_idx = i
                break

    if header_idx < 0:
        return title, [[line] for line in lines[1 : max_rows + 1]]

    header_tokens: list[str] = []
    idx = header_idx
    while idx < len(lines):
        header_tokens.append(lines[idx])
        idx += 1
        if lines[idx - 1] in {"岗位要求", "要求", "条件"}:
            break
        if len(header_tokens) >= 6:
            break

    has_dept = header_tokens[0] == "部门"
    rows: list[list[str]] = []
    current_dept = ""
    data = lines[idx:]
    i = 0
    while i < len(data) and len(rows) < max_rows:
        if has_dept:
            if i + 4 >= len(data):
                break
            dept = data[i]
            job = data[i + 1]
            major = data[i + 2]
            count = data[i + 3]
            requirement = data[i + 4]
            i += 5
            if len(dept) <= 8 and not re.search(r"\d", dept):
                current_dept = dept
            else:
                job, major, count, requirement = dept, job, major, count
            rows.append([current_dept, job, major, count, requirement])
        else:
            if i + 3 >= len(data):
                break
            job, major, count, requirement = data[i : i + 4]
            rows.append([job, major, count, requirement])
            i += 4

    if not rows:
        return title, [[line] for line in lines[1 : max_rows + 1]]

    headers = ["部门", "岗位", "专业", "人数", "岗位要求"] if has_dept else ["岗位", "专业", "人数", "岗位要求"]
    return title, [headers, *rows]


def _word_tables(local_path: Path, max_rows: int) -> list[list[list[str]]]:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    word = None
    doc = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        doc = word.Documents.Open(str(local_path.resolve()), ReadOnly=True, ConfirmConversions=False)
        tables: list[list[list[str]]] = []
        table_count = min(int(doc.Tables.Count), 3)
        for table_index in range(1, table_count + 1):
            table = doc.Tables(table_index)
            rows: list[list[str]] = []
            row_count = min(int(table.Rows.Count), max_rows + 1)
            col_count = min(int(table.Columns.Count), 8)
            for r in range(1, row_count + 1):
                cells: list[str] = []
                for c in range(1, col_count + 1):
                    try:
                        value = str(table.Cell(r, c).Range.Text or "")
                    except Exception:
                        value = ""
                    value = value.replace("\r", "").replace("\a", "")
                    cells.append(_clean_line(value))
                if any(cells):
                    rows.append(cells)
            if rows:
                tables.append(rows)
        return tables
    finally:
        try:
            if doc is not None:
                doc.Close(False)
        finally:
            if word is not None:
                word.Quit()
            pythoncom.CoUninitialize()


def _excel_tables(local_path: Path, max_rows: int) -> list[list[list[str]]]:
    suffix = local_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(local_path, read_only=True, data_only=True)
        tables: list[list[list[str]]] = []
        for sheet in wb.worksheets[:3]:
            rows: list[list[str]] = []
            for row in sheet.iter_rows(min_row=1, max_row=max_rows + 1, values_only=True):
                cells = [_clean_line(cell) for cell in row]
                while cells and not cells[-1]:
                    cells.pop()
                if any(cells):
                    rows.append(cells[:8])
            if rows:
                tables.append(rows)
        return tables

    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(local_path.resolve()), ReadOnly=True)
        tables: list[list[list[str]]] = []
        sheet_count = min(int(workbook.Worksheets.Count), 3)
        for sheet_index in range(1, sheet_count + 1):
            sheet = workbook.Worksheets(sheet_index)
            used = sheet.UsedRange
            values = used.Value
            rows: list[list[str]] = []
            if values is None:
                continue
            if not isinstance(values, tuple):
                values = ((values,),)
            elif values and not isinstance(values[0], tuple):
                values = (values,)
            for row in list(values)[: max_rows + 1]:
                cells = [_clean_line(cell) for cell in row]
                while cells and not cells[-1]:
                    cells.pop()
                if any(cells):
                    rows.append(cells[:8])
            if rows:
                tables.append(rows)
        return tables
    finally:
        try:
            if workbook is not None:
                workbook.Close(False)
        finally:
            if excel is not None:
                excel.Quit()
            pythoncom.CoUninitialize()


def _tables_from_file(local_path: str, max_rows: int) -> list[list[list[str]]]:
    path = Path(local_path or "")
    if not path.exists():
        return []
    suffix = path.suffix.lower()
    try:
        if suffix in {".doc", ".docx"}:
            return _word_tables(path, max_rows)
        if suffix in {".xls", ".xlsx", ".xlsm"}:
            return _excel_tables(path, max_rows)
    except Exception as exc:
        print(f"[attachment-images] structured table extraction failed: {path} {exc}", flush=True)
    return []


def _is_job_table_candidate(name: str, parsed_text: str, local_path: str) -> bool:
    haystack = f"{name}\n{parsed_text}\n{local_path}"
    excluded = ["承诺书", "申请表", "报名表", "登记表", "诚信", "身份证", "照片"]
    if any(word in haystack for word in excluded):
        return False
    positive = ["岗位", "职位", "招聘计划", "岗位信息", "岗位要求", "招聘人数", "引进计划数", "专业要求"]
    return any(word in haystack for word in positive)


def _render_table_image(title: str, rows: list[list[str]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    title_font = _font(34)
    header_font = _font(24)
    body_font = _font(22)
    note_font = _font(18)
    width = 1180
    margin = 34
    table_width = width - margin * 2
    if not rows:
        rows = [["暂无可展示的岗位表信息"]]

    col_count = max(len(row) for row in rows)
    if col_count >= 5:
        col_widths = [140, 210, 300, 90, table_width - 140 - 210 - 300 - 90]
        if col_count > 5:
            rows = [row[:4] + ["；".join(row[4:])] for row in rows]
            col_count = 5
    elif col_count == 4:
        col_widths = [220, 330, 100, table_width - 220 - 330 - 100]
    else:
        col_widths = [table_width]
        col_count = 1

    probe = Image.new("RGB", (width, 100), "white")
    draw = ImageDraw.Draw(probe)
    rendered_rows: list[list[list[str]]] = []
    row_heights: list[int] = []
    for row_index, row in enumerate(rows):
        font = header_font if row_index == 0 and col_count > 1 else body_font
        rendered_cells: list[list[str]] = []
        max_lines = 1
        for col_index in range(col_count):
            value = row[col_index] if col_index < len(row) else ""
            wrapped = _wrap_text(draw, value, font, col_widths[col_index] - 18)
            rendered_cells.append(wrapped)
            max_lines = max(max_lines, len(wrapped))
        rendered_rows.append(rendered_cells)
        row_heights.append(max(48, max_lines * 30 + 20))

    title_lines = _wrap_text(draw, title, title_font, table_width)
    height = margin + len(title_lines) * 44 + 18 + sum(row_heights) + margin + 34
    image = Image.new("RGB", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)

    y = margin
    for line in title_lines:
        draw.text((margin, y), line, fill="#111827", font=title_font)
        y += 44
    y += 14

    x_positions = [margin]
    for col_width in col_widths[:-1]:
        x_positions.append(x_positions[-1] + col_width)

    for row_index, cells in enumerate(rendered_rows):
        row_h = row_heights[row_index]
        bg = "#e8f3ee" if row_index == 0 and col_count > 1 else ("#ffffff" if row_index % 2 else "#f8fafc")
        draw.rectangle((margin, y, margin + table_width, y + row_h), fill=bg, outline="#d1d5db")
        for col_index, cell_lines in enumerate(cells):
            x = x_positions[col_index]
            col_w = col_widths[col_index]
            draw.line((x, y, x, y + row_h), fill="#d1d5db", width=1)
            font = header_font if row_index == 0 and col_count > 1 else body_font
            text_y = y + 10
            for line in cell_lines:
                draw.text((x + 9, text_y), line, fill="#111827", font=font)
                text_y += 30
        draw.line((margin + table_width, y, margin + table_width, y + row_h), fill="#d1d5db", width=1)
        y += row_h

    y += 10
    draw.text((margin, y), "注：图片由公告附件解析生成，请以文末原文链接及官方附件为准。", fill="#6b7280", font=note_font)
    image.save(output_path)


def generate_attachment_table_images(source_id: str, *, max_images: int = 2, max_rows: int = 14) -> list[Path]:
    with _connect() as conn:
        records = conn.execute(
            """
            SELECT id, name, parsed_text, local_path
            FROM gongkao_event_attachments
            WHERE event_source_id = ?
              AND (
                  coalesce(parsed_text, '') <> ''
                  OR coalesce(local_path, '') <> ''
              )
            ORDER BY
              CASE WHEN coalesce(local_path, '') <> '' THEN 0 ELSE 1 END,
              id ASC
            LIMIT ?
            """,
            (source_id, max_images),
        ).fetchall()

    paths: list[Path] = []
    for record in records:
        name = str(record["name"] or "")
        parsed_text = str(record["parsed_text"] or "")
        local_path = str(record["local_path"] or "")
        if not _is_job_table_candidate(name, parsed_text, local_path):
            continue
        tables = _tables_from_file(str(record["local_path"] or ""), max_rows=max_rows)
        if tables:
            rows = tables[0]
        else:
            _, rows = _rows_from_parsed_text(parsed_text, max_rows=max_rows)
        if not rows:
            continue
        safe_id = re.sub(r"[^A-Za-z0-9_-]+", "_", str(source_id))
        output_path = OUTPUT_ROOT / safe_id / f"attachment_table_{record['id']}.png"
        _render_table_image(name or "附件岗位表", rows, output_path)
        paths.append(output_path)
    return paths
