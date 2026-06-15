from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import mimetypes
import re
import sqlite3
import time
import warnings
import zipfile
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urljoin, urlparse
from xml.etree import ElementTree as ET

warnings.filterwarnings("ignore", message=r".*urllib3.*chardet.*charset_normalizer.*")
import requests

from .config import CONFIG
from .schema import init_db


ATTACHMENT_ROOT = CONFIG.project_root / "attachments"
ATTACHMENT_EXTS = {
    "doc",
    "docx",
    "xls",
    "xlsx",
    "xlsm",
    "csv",
    "pdf",
    "txt",
    "zip",
    "rar",
    "wps",
}

JOB_TABLE_INCLUDE_WORDS = (
    "岗位表",
    "职位表",
    "岗位计划",
    "招聘计划",
    "岗位信息",
    "岗位需求",
    "需求表",
    "岗位一览",
    "职位一览",
    "条件表",
    "岗位及条件",
)

JOB_TABLE_EXCLUDE_WORDS = (
    "报名表",
    "申请表",
    "承诺书",
    "诚信",
    "专业目录",
    "操作说明",
    "操作流程",
    "考试大纲",
    "须知",
    "证明",
    "资格审查",
    "信息采集",
    "知情书",
)


@dataclass
class EventRow:
    source_platform: str
    source_id: str
    title: str
    source_url: str
    article_url: str
    source_origin_url: str
    source_origin_html: str
    raw_json: str


def _connect(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def _clean_text(value: Any) -> str:
    text = html.unescape(str(value or ""))
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _safe_filename(value: str, fallback: str) -> str:
    name = unquote(value or "").strip().replace("\\", "_").replace("/", "_")
    name = re.sub(r'[<>:"|?*\x00-\x1f]', "_", name)
    name = re.sub(r"\s+", " ", name).strip(" .")
    return name[:120] or fallback


def _guess_ext(url: str, name: str, content_type: str = "") -> str:
    for candidate in (name, urlparse(url).path):
        match = re.search(r"\.([A-Za-z0-9]{2,5})(?:$|[?#])", candidate or "")
        if match:
            ext = match.group(1).lower()
            if ext in ATTACHMENT_EXTS:
                return ext
    guessed = mimetypes.guess_extension((content_type or "").split(";")[0].strip())
    return (guessed or "").lstrip(".").lower()


def _looks_like_attachment(url: str, text: str) -> bool:
    lower = f"{url} {text}".lower()
    if any(f".{ext}" in lower for ext in ATTACHMENT_EXTS):
        return True
    return any(word in text for word in ["附件", "岗位表", "职位表", "报名表", "下载", "招聘计划", "人员名单"])


def is_job_table_candidate(name: str, href: str = "") -> bool:
    decoded_href = unquote(href or "")
    href_lower = decoded_href.lower()
    if "fenbi.com/api/website/article/crawler/check" in href_lower:
        return False
    if "hera-webapp.fenbi.com/api/website/article/crawler/check" in href_lower:
        return False
    text = f"{name or ''} {decoded_href}"
    if any(word in text for word in JOB_TABLE_EXCLUDE_WORDS):
        return False
    return any(word in text for word in JOB_TABLE_INCLUDE_WORDS)


class AttachmentHtmlParser(HTMLParser):
    def __init__(self, base_url: str) -> None:
        super().__init__()
        self.base_url = base_url
        self.items: list[dict[str, str]] = []
        self._anchor: dict[str, str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = {key.lower(): value or "" for key, value in attrs}
        if tag.lower() == "a":
            href = attrs_dict.get("href", "")
            oldsrc = attrs_dict.get("oldsrc", "") or attrs_dict.get("data-url", "") or attrs_dict.get("data-href", "")
            self._anchor = {"href": href, "oldsrc": oldsrc, "text": ""}

    def handle_data(self, data: str) -> None:
        if self._anchor is not None:
            self._anchor["text"] += data or ""

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() != "a" or self._anchor is None:
            return
        href = self._anchor.get("href", "").strip()
        oldsrc = self._anchor.get("oldsrc", "").strip()
        text = _clean_text(self._anchor.get("text", ""))
        raw_url = oldsrc or href
        if raw_url:
            full_url = urljoin(self.base_url, raw_url)
            if _looks_like_attachment(full_url, text):
                self.items.append({"name": text, "href": full_url, "oldsrc": oldsrc})
        self._anchor = None


def extract_attachment_links(source_html: str, base_url: str, raw_json: str = "") -> list[dict[str, str]]:
    parser = AttachmentHtmlParser(base_url)
    parser.feed(source_html or "")
    items = list(parser.items)
    try:
        raw = json.loads(raw_json or "{}")
    except Exception:
        raw = {}
    for key in ("article_attachments", "origin_attachments"):
        value = raw.get(key) if isinstance(raw, dict) else None
        if not isinstance(value, list):
            continue
        for item in value:
            if not isinstance(item, dict):
                continue
            href = str(item.get("href") or item.get("url") or "").strip()
            oldsrc = str(item.get("oldsrc") or "").strip()
            name = _clean_text(item.get("text") or item.get("name") or "")
            raw_url = oldsrc or href
            if raw_url:
                full_url = urljoin(base_url, raw_url)
                if _looks_like_attachment(full_url, name):
                    items.append({"name": name, "href": full_url, "oldsrc": oldsrc})

    deduped: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in items:
        href = item.get("href", "")
        key = href.split("#", 1)[0]
        if not href or key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def _content_disposition_filename(value: str) -> str:
    if not value:
        return ""
    match = re.search(r"filename\*=UTF-8''([^;]+)", value, flags=re.I)
    if match:
        return unquote(match.group(1))
    match = re.search(r'filename="?([^";]+)"?', value, flags=re.I)
    return unquote(match.group(1)) if match else ""


def _download(url: str, referer: str = "") -> tuple[bytes, str, str]:
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "*/*",
    }
    if referer:
        headers["Referer"] = referer
    response = requests.get(url, headers=headers, timeout=45, allow_redirects=True)
    response.raise_for_status()
    content_type = response.headers.get("Content-Type", "")
    filename = _content_disposition_filename(response.headers.get("Content-Disposition", ""))
    return response.content, content_type, filename


def _parse_xlsx(content: bytes) -> tuple[str, dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    text_rows: list[str] = []
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets[:5]:
        sheet_rows: list[list[str]] = []
        text_rows.append(f"[sheet] {sheet.title}")
        for row in sheet.iter_rows(min_row=1, max_row=80, values_only=True):
            cells = [_clean_text(cell) for cell in row if _clean_text(cell)]
            if not cells:
                continue
            sheet_rows.append(cells)
            text_rows.append(" | ".join(cells))
        sheets.append({"name": sheet.title, "rows": sheet_rows[:80]})
    return "\n".join(text_rows), {"sheets": sheets}


def _parse_docx(content: bytes) -> tuple[str, dict[str, Any]]:
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        joined = _clean_text("".join(texts))
        if joined:
            paragraphs.append(joined)
    return "\n".join(paragraphs), {"paragraphs": paragraphs[:300]}


def _parse_csv_or_text(content: bytes) -> tuple[str, dict[str, Any]]:
    decoded = ""
    used_encoding = ""
    for encoding in ("utf-8-sig", "gb18030", "gbk", "utf-8"):
        try:
            decoded = content.decode(encoding)
            used_encoding = encoding
            break
        except Exception:
            continue
    if not decoded:
        decoded = content.decode("utf-8", errors="ignore")
        used_encoding = "utf-8-ignore"
    rows = list(csv.reader(io.StringIO(decoded)))[:80]
    return decoded[:20000], {"encoding": used_encoding, "rows": rows}


def _parse_doc_with_word(local_path: Path) -> tuple[str, dict[str, Any]]:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    word = None
    doc = None
    try:
        word = win32com.client.DispatchEx("Word.Application")
        word.Visible = False
        doc = word.Documents.Open(str(local_path.resolve()), ReadOnly=True, ConfirmConversions=False)
        text = str(doc.Content.Text or "")
        lines = [_clean_text(line) for line in text.splitlines()]
        lines = [line for line in lines if line]
        return "\n".join(lines), {"lines": lines[:500]}
    finally:
        try:
            if doc is not None:
                doc.Close(False)
        finally:
            if word is not None:
                word.Quit()
            pythoncom.CoUninitialize()


def _parse_xls_with_excel(local_path: Path) -> tuple[str, dict[str, Any]]:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(local_path.resolve()), ReadOnly=True)
        text_rows: list[str] = []
        sheets: list[dict[str, Any]] = []
        sheet_count = min(int(workbook.Worksheets.Count), 5)
        for sheet_index in range(1, sheet_count + 1):
            sheet = workbook.Worksheets(sheet_index)
            used = sheet.UsedRange
            values = used.Value
            rows: list[list[str]] = []
            text_rows.append(f"[sheet] {sheet.Name}")
            if values is None:
                continue
            if not isinstance(values, tuple):
                values = ((values,),)
            elif values and not isinstance(values[0], tuple):
                values = (values,)
            for row in list(values)[:80]:
                cells = [_clean_text(cell) for cell in row if _clean_text(cell)]
                if cells:
                    rows.append(cells)
                    text_rows.append(" | ".join(cells))
            sheets.append({"name": str(sheet.Name), "rows": rows[:80]})
        return "\n".join(text_rows), {"sheets": sheets}
    finally:
        try:
            if workbook is not None:
                workbook.Close(False)
        finally:
            if excel is not None:
                excel.Quit()
            pythoncom.CoUninitialize()


def _is_invalid_download(content: bytes, ext: str, content_type: str) -> str:
    stripped = content.strip().lower()
    if ext in {"doc", "docx", "xls", "xlsx", "xlsm", "pdf"} and len(content) < 128:
        return f"下载内容过小，不像有效 .{ext} 文件。"
    if stripped in {b"success", b"ok", b"true"}:
        return "下载得到的是接口状态文本，不是实际附件文件。"
    if ext == "pdf" and not content.startswith(b"%PDF"):
        return "下载内容不是有效 PDF。"
    if ext in {"doc", "xls"} and not content.startswith(b"\xd0\xcf\x11\xe0"):
        return f"下载内容不是有效旧版 .{ext} 文件。"
    if ext in {"docx", "xlsx", "xlsm"} and not content.startswith(b"PK"):
        return f"下载内容不是有效 .{ext} 文件。"
    return ""


def parse_attachment(
    content: bytes,
    ext: str,
    local_path: Path | None = None,
    *,
    use_office_com: bool = False,
) -> tuple[str, str, dict[str, Any], str]:
    try:
        if ext in {"xlsx", "xlsm"}:
            text, data = _parse_xlsx(content)
            return "parsed", text[:30000], data, ""
        if ext == "xls" and local_path is not None and use_office_com:
            text, data = _parse_xls_with_excel(local_path)
            return "parsed", text[:30000], data, ""
        if ext == "docx":
            text, data = _parse_docx(content)
            return "parsed", text[:30000], data, ""
        if ext == "doc" and local_path is not None and use_office_com:
            text, data = _parse_doc_with_word(local_path)
            return "parsed", text[:30000], data, ""
        if ext in {"csv", "txt", "html", "htm"}:
            text, data = _parse_csv_or_text(content)
            return "parsed", text[:30000], data, ""
        if ext == "xls":
            return "downloaded_unparsed", "", {}, "旧版 .xls 暂未解析，已下载到本地。"
        if ext == "doc":
            return "downloaded_unparsed", "", {}, "旧版 .doc 暂未解析，已下载到本地。"
        if ext == "pdf":
            return "downloaded_unparsed", "", {}, "PDF 暂未启用文本解析，已下载到本地。"
        if ext in {"zip", "rar", "wps"}:
            return "downloaded_unparsed", "", {}, f"{ext} 暂未解析，已下载到本地。"
        return "downloaded_unparsed", "", {}, "未知附件类型，已下载到本地。"
    except Exception as exc:
        return "parse_failed", "", {}, str(exc)


def _event_dir(source_id: str) -> Path:
    safe_id = _safe_filename(source_id, "event")
    path = ATTACHMENT_ROOT / safe_id
    path.mkdir(parents=True, exist_ok=True)
    return path


def _save_file(source_id: str, url: str, name: str, content: bytes, content_type: str, server_filename: str) -> tuple[Path, str]:
    ext = _guess_ext(url, server_filename or name, content_type)
    fallback = hashlib.sha1(url.encode("utf-8")).hexdigest()[:12]
    filename = _safe_filename(server_filename or name or Path(urlparse(url).path).name, fallback)
    if ext and not filename.lower().endswith(f".{ext}"):
        filename = f"{filename}.{ext}"
    target = _event_dir(source_id) / filename
    if target.exists():
        stem = target.stem
        suffix = target.suffix
        digest = hashlib.sha1(url.encode("utf-8")).hexdigest()[:8]
        target = target.with_name(f"{stem}_{digest}{suffix}")
    target.write_bytes(content)
    return target, ext


def _fetch_events(db_path: Path, topic_ids: list[str], limit: int, only_missing: bool) -> list[EventRow]:
    clauses = ["coalesce(source_origin_html, '') <> ''"]
    values: list[Any] = []
    if topic_ids:
        placeholders = ",".join("?" for _ in topic_ids)
        clauses.append(f"source_id IN ({placeholders})")
        values.extend(topic_ids)
    if only_missing:
        clauses.append(
            """
            NOT EXISTS (
                SELECT 1 FROM gongkao_event_attachments a
                WHERE a.event_source_platform = gongkao_events.source_platform
                  AND a.event_source_id = gongkao_events.source_id
                  AND coalesce(a.download_status, '') = 'downloaded'
            )
            """
        )
    query = f"""
        SELECT source_platform, source_id, title, source_url, article_url,
               source_origin_url, source_origin_html, raw_json
        FROM gongkao_events
        WHERE {' AND '.join(clauses)}
        ORDER BY imported_at DESC
        LIMIT ?
    """
    values.append(limit)
    with _connect(db_path) as conn:
        rows = conn.execute(query, values).fetchall()
    return [EventRow(**dict(row)) for row in rows]


def _upsert_attachment(conn: sqlite3.Connection, event: EventRow, record: dict[str, Any]) -> None:
    imported_at = datetime.utcnow().isoformat()
    conn.execute(
        """
        INSERT INTO gongkao_event_attachments(
            event_source_platform, event_source_id, attachment_scope, name,
            href, oldsrc, file_ext, download_status, local_path, file_size,
            content_type, parse_status, parsed_text, parsed_json, error_message,
            raw_json, imported_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(event_source_platform, event_source_id, attachment_scope, oldsrc, href, name)
        DO UPDATE SET
            file_ext=excluded.file_ext,
            download_status=CASE
                WHEN excluded.download_status = 'pending'
                 AND coalesce(gongkao_event_attachments.download_status, '') = 'downloaded'
                THEN gongkao_event_attachments.download_status
                ELSE excluded.download_status
            END,
            local_path=CASE
                WHEN excluded.download_status = 'pending'
                 AND coalesce(gongkao_event_attachments.local_path, '') <> ''
                THEN gongkao_event_attachments.local_path
                ELSE excluded.local_path
            END,
            file_size=CASE
                WHEN excluded.download_status = 'pending'
                 AND coalesce(gongkao_event_attachments.file_size, 0) > 0
                THEN gongkao_event_attachments.file_size
                ELSE excluded.file_size
            END,
            content_type=CASE
                WHEN excluded.download_status = 'pending'
                 AND coalesce(gongkao_event_attachments.content_type, '') <> ''
                THEN gongkao_event_attachments.content_type
                ELSE excluded.content_type
            END,
            parse_status=CASE
                WHEN excluded.download_status = 'pending'
                 AND coalesce(gongkao_event_attachments.parse_status, '') <> ''
                THEN gongkao_event_attachments.parse_status
                WHEN coalesce(excluded.parsed_text, '') <> '' THEN excluded.parse_status
                WHEN coalesce(gongkao_event_attachments.parsed_text, '') <> '' THEN gongkao_event_attachments.parse_status
                ELSE excluded.parse_status
            END,
            parsed_text=CASE
                WHEN coalesce(excluded.parsed_text, '') <> '' THEN excluded.parsed_text
                ELSE gongkao_event_attachments.parsed_text
            END,
            parsed_json=CASE
                WHEN coalesce(excluded.parsed_text, '') <> '' THEN excluded.parsed_json
                ELSE gongkao_event_attachments.parsed_json
            END,
            error_message=CASE
                WHEN excluded.download_status = 'pending'
                 AND coalesce(gongkao_event_attachments.error_message, '') <> ''
                THEN gongkao_event_attachments.error_message
                WHEN coalesce(excluded.parsed_text, '') <> '' THEN excluded.error_message
                WHEN coalesce(gongkao_event_attachments.parsed_text, '') <> '' THEN gongkao_event_attachments.error_message
                ELSE excluded.error_message
            END,
            raw_json=excluded.raw_json,
            imported_at=excluded.imported_at
        """,
        (
            event.source_platform,
            event.source_id,
            record.get("attachment_scope", "origin"),
            record.get("name", ""),
            record.get("href", ""),
            record.get("oldsrc", ""),
            record.get("file_ext", ""),
            record.get("download_status", "pending"),
            record.get("local_path", ""),
            record.get("file_size", 0),
            record.get("content_type", ""),
            record.get("parse_status", "metadata_only"),
            record.get("parsed_text", ""),
            record.get("parsed_json", "{}"),
            record.get("error_message", ""),
            record.get("raw_json", "{}"),
            imported_at,
        ),
    )


def process_event(
    db_path: Path,
    event: EventRow,
    max_attachments: int,
    *,
    use_office_com: bool,
    metadata_only: bool = False,
    job_tables_only: bool = False,
) -> tuple[int, int, int, int]:
    base_url = event.source_origin_url or event.article_url or event.source_url
    links = extract_attachment_links(event.source_origin_html, base_url, event.raw_json)
    if job_tables_only:
        links = [link for link in links if is_job_table_candidate(link.get("name", ""), link.get("href", ""))]
    if max_attachments > 0:
        links = links[:max_attachments]
    registered = 0
    downloaded = 0
    parsed = 0
    failed = 0
    mode = "仅登记" if metadata_only else ("仅岗位表下载解析" if job_tables_only else "下载解析")
    print(f"[attachments] {event.source_id} {event.title} 发现附件 {len(links)} 个，模式：{mode}", flush=True)
    with _connect(db_path) as conn:
        for index, link in enumerate(links, 1):
            url = link.get("href", "")
            name = link.get("name", "") or Path(urlparse(url).path).name or f"附件{index}"
            record: dict[str, Any] = {
                "attachment_scope": "origin",
                "name": name,
                "href": url,
                "oldsrc": link.get("oldsrc", ""),
                "raw_json": json.dumps(link, ensure_ascii=False),
                "file_ext": _guess_ext(url, name),
            }
            if metadata_only:
                record.update(
                    {
                        "download_status": "pending",
                        "parse_status": "metadata_only",
                    }
                )
                _upsert_attachment(conn, event, record)
                conn.commit()
                registered += 1
                if index % 20 == 0 or index == len(links):
                    print(f"[attachments] 已登记 {index}/{len(links)} 个附件链接", flush=True)
                continue
            try:
                print(f"[attachments] 下载 {index}/{len(links)}: {name} {url}", flush=True)
                content, content_type, server_filename = _download(url, referer=base_url)
                ext = _guess_ext(url, server_filename or name, content_type)
                invalid_reason = _is_invalid_download(content, ext, content_type)
                if invalid_reason:
                    raise RuntimeError(invalid_reason)
                local_path, ext = _save_file(event.source_id, url, name, content, content_type, server_filename)
                parse_status, parsed_text, parsed_json, parse_error = parse_attachment(
                    content,
                    ext,
                    local_path,
                    use_office_com=use_office_com,
                )
                downloaded += 1
                if parse_status == "parsed":
                    parsed += 1
                if parse_status == "parse_failed":
                    failed += 1
                record.update(
                    {
                        "file_ext": ext,
                        "download_status": "downloaded",
                        "local_path": str(local_path),
                        "file_size": len(content),
                        "content_type": content_type,
                        "parse_status": parse_status,
                        "parsed_text": parsed_text,
                        "parsed_json": json.dumps(parsed_json, ensure_ascii=False),
                        "error_message": parse_error,
                    }
                )
                print(f"[attachments] 保存: {local_path}，解析状态: {parse_status}", flush=True)
            except Exception as exc:
                failed += 1
                ext = _guess_ext(url, name)
                record.update(
                    {
                        "file_ext": ext,
                        "download_status": "failed",
                        "parse_status": "download_failed",
                        "error_message": str(exc),
                    }
                )
                print(f"[attachments] 失败: {name}，原因: {exc}", flush=True)
            _upsert_attachment(conn, event, record)
            conn.commit()
            registered += 1
    return registered, downloaded, parsed, failed


def main() -> None:
    parser = argparse.ArgumentParser(description="Download and parse gongkao announcement attachments.")
    parser.add_argument("--db", default=str(CONFIG.database_path))
    parser.add_argument("--topic_id", default="", help="Single event source_id.")
    parser.add_argument("--topic_ids", default="", help="Comma-separated source_id list.")
    parser.add_argument("--limit", type=int, default=20)
    parser.add_argument("--max_attachments", type=int, default=10)
    parser.add_argument("--only_missing", action="store_true")
    parser.add_argument("--metadata_only", action="store_true", help="Only scan and save attachment link metadata; do not download files.")
    parser.add_argument("--job_tables_only", action="store_true", help="Only download/parse likely job table attachments.")
    parser.add_argument(
        "--use_office_com",
        action="store_true",
        help="Use local Microsoft Office COM to parse legacy .doc/.xls files. Best used from an interactive terminal.",
    )
    args = parser.parse_args()

    db_path = Path(args.db)
    init_db(db_path)
    topic_ids = [item.strip() for item in args.topic_ids.split(",") if item.strip()]
    if args.topic_id.strip():
        topic_ids.insert(0, args.topic_id.strip())
    topic_ids = list(dict.fromkeys(topic_ids))

    events = _fetch_events(db_path, topic_ids, args.limit, args.only_missing)
    if not events:
        print("[attachments] 没有找到可处理的公告。", flush=True)
        return

    total_downloaded = 0
    total_parsed = 0
    total_failed = 0
    total_registered = 0
    for i, event in enumerate(events, 1):
        print("=" * 72, flush=True)
        print(f"[attachments] 进度 {i}/{len(events)}", flush=True)
        registered, downloaded, parsed, failed = process_event(
            db_path,
            event,
            args.max_attachments,
            use_office_com=args.use_office_com,
            metadata_only=args.metadata_only,
            job_tables_only=args.job_tables_only,
        )
        total_registered += registered
        total_downloaded += downloaded
        total_parsed += parsed
        total_failed += failed
        if not args.metadata_only:
            time.sleep(0.2)

    print("=" * 72, flush=True)
    print(
        f"[attachments] 处理完成：登记 {total_registered} 个，下载成功 {total_downloaded} 个，解析成功 {total_parsed} 个，失败 {total_failed} 个。",
        flush=True,
    )
    print(f"[attachments] 本地目录: {ATTACHMENT_ROOT}", flush=True)


if __name__ == "__main__":
    main()
