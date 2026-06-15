from __future__ import annotations

import argparse
import asyncio
import hashlib
import io
import json
import re
import urllib.parse
import zipfile
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import requests
from playwright.async_api import Browser, Page, async_playwright

from .config import CONFIG
from .store import ContentStore


BASE_URL = "https://www.gongkaoleida.com"
CATEGORY_PATHS = {
    "all": "/exam_search",
    "gwy": "/exam_search/1-2",
    "gongwuyuan": "/exam_search/1-2",
    "sydw": "/exam_search/1-3",
    "shiye": "/exam_search/1-3",
    "guoqi": "/exam_search/1-78",
    "teacher": "/exam_search/1-59",
    "medical": "/exam_search/1-60",
    "xuandiao": "/exam_search/1-7",
}


@dataclass
class GongkaoEvent:
    source_platform: str
    source_id: str
    title: str
    region: str
    category: str
    org_name: str
    job_count: int | None
    qualification: str
    major_requirements: str
    registration_start: str
    registration_deadline: str
    exam_date: str
    status: str
    publish_time: str
    source_url: str
    article_url: str
    summary: str
    raw_text: str
    raw_json: str
    hash_id: str
    source_origin_url: str
    source_origin_text: str
    source_origin_html: str
    origin_search_status: str
    origin_search_attempts: int
    origin_last_checked_at: str


def _clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _normalize_date(date_text: str, default_year: int | None = None) -> str:
    text = _clean_text(date_text)
    if not text:
        return ""
    m = re.search(r"(?:(\d{4})[年./-])?\s*(\d{1,2})[月./-](\d{1,2})日?", text)
    if not m:
        return ""
    year = int(m.group(1) or default_year or datetime.now().year)
    month = int(m.group(2))
    day = int(m.group(3))
    return f"{year:04d}-{month:02d}-{day:02d}"


def _extract_between(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.S)
        if m:
            return _clean_text(m.group(1))
    return ""


def _extract_first_date(text: str, default_year: int | None = None) -> str:
    matches = re.findall(r"((?:(?:\d{4})[年./-])?\s*\d{1,2}[月./-]\d{1,2}日?)", text)
    for match in matches:
        normalized = _normalize_date(match, default_year)
        if normalized:
            return normalized
    return ""


def _extract_registration_dates(raw_text: str, publish_time: str) -> tuple[str, str]:
    year = None
    if publish_time:
        try:
            year = datetime.fromisoformat(publish_time[:19]).year
        except Exception:
            year = None

    start = _extract_between(
        raw_text,
        [
            r"报名时间[：:]\s*([0-9年月日:\-至—~～\s]+)",
            r"报名(?:起止|期限|期间)[：:]\s*([0-9年月日:\-至—~～\s]+)",
            r"报名时间[：:]\s*([^\n。；]+)",
            r"报名(?:方式)?[：:].{0,40}?(\d{4}年\d{1,2}月\d{1,2}日[^\n。；]*)",
            r"于([0-9年月日:\-至—\s]+?)报名截止",
        ],
    )
    deadline = _extract_between(
        raw_text,
        [
            r"报名截止(?:日期|时间)?[：:]\s*([0-9年月日:\-至—~～\s]+)",
            r"截止报名(?:日期|时间)?[：:]\s*([0-9年月日:\-至—~～\s]+)",
            r"于\s*([0-9年月日:\-至—~～\s]+?)前",
            r"报名截止(?:日期|时间)?[：:]\s*([^\n。；]+)",
            r"截止(?:日期|时间)?[：:]\s*([^\n。；]+)",
            r"报名时间[：:][^\n。；]*?至\s*([^\n。；]+)",
            r"报名(?:时间|期限|期间)[：:]\s*[^\n。；]*?[至到—-]\s*([0-9年月日:\-:\s]+)",
        ],
    )

    start_date = _normalize_date(start, year)
    deadline_date = _normalize_date(deadline, year)
    if not start_date and start:
        start_date = _extract_first_date(start, year)
    if not deadline_date and deadline:
        deadline_date = _extract_first_date(deadline, year)
    if (not start_date or not deadline_date) and start and any(sep in start for sep in ["至", "到", "-", "—", "~", "～"]):
        parts = re.split(r"\s*(?:至|到|—|-|~|～)\s*", start)
        if len(parts) >= 2:
            start_date = start_date or _extract_first_date(parts[0], year)
            deadline_date = deadline_date or _extract_first_date(parts[-1], year)
    return start_date, deadline_date


def _extract_exam_date(raw_text: str, publish_time: str) -> str:
    year = None
    if publish_time:
        try:
            year = datetime.fromisoformat(publish_time[:19]).year
        except Exception:
            year = None
    exam_line = _extract_between(
        raw_text,
        [
            r"笔试(?:时间)?[：:]\s*([^\n。；]+)",
            r"考试(?:时间)?[：:]\s*([^\n。；]+)",
            r"面试(?:时间)?[：:]\s*([^\n。；]+)",
        ],
    )
    return _normalize_date(exam_line, year)


def _extract_qualification(raw_text: str) -> str:
    matches = re.findall(r"(博士研究生|硕士研究生|硕士|本科及以上|本科|专科及以上|专科|大专及以上|中专及以上|高中及以上)", raw_text)
    deduped: list[str] = []
    for item in matches:
        if item not in deduped:
            deduped.append(item)
    return "、".join(deduped)


def _extract_major_requirements(raw_text: str) -> str:
    keywords = ["计算机", "软件工程", "信息安全", "网络工程", "电子信息", "不限专业", "专业不限", "数据", "人工智能"]
    hits = [kw for kw in keywords if kw in raw_text]
    return "、".join(dict.fromkeys(hits))


def _looks_like_attachment_url(url: str, text: str = "") -> bool:
    lower_url = url.lower()
    if "javascript:;" in lower_url:
        return "附件" in text or "岗位表" in text or "报名表" in text or "需求表" in text
    return any(ext in lower_url for ext in [".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt", ".zip", ".rar"]) or any(
        keyword in text for keyword in ["附件", "岗位表", "职位表", "报名表", "需求表", "招考简章"]
    )


def _normalize_title_for_match(title: str) -> str:
    text = re.sub(r"20\d{2}", " ", title)
    text = re.sub(r"(公告|招聘|招录|公开招聘|面向社会公开招聘|人员招聘|招聘公告|上半年|下半年)", " ", text)
    return _clean_text(text)


def _title_chunks(title: str) -> list[str]:
    normalized = _normalize_title_for_match(title)
    parts = re.split(r"[，,。；：:（）()【】\[\]\s/\-]+", normalized)
    stopwords = {"国家", "中国", "公开", "社会", "人员", "事业单位"}
    chunks = [part.strip() for part in parts if len(part.strip()) >= 3 and part.strip() not in stopwords]
    return chunks


def _is_origin_text_match(title: str, text: str, url: str) -> bool:
    haystack = f"{text} {url}"
    chunks = _title_chunks(title)
    if not chunks:
        return False
    hits = [chunk for chunk in chunks if chunk in haystack]
    if len(hits) >= 2:
        return True
    if len(chunks) == 1 and len(hits) == 1:
        return True
    # The longest chunk is usually the organization core; keep it as a strict fallback.
    longest = max(chunks, key=len)
    return len(longest) >= 6 and longest in haystack


def _extract_urls(text: str) -> list[str]:
    return re.findall(r"https?://[^\s\"'<>）)]+", text)


def _clean_url(url: str) -> str:
    return url.strip().rstrip(".,;，。；")


def _extract_origin_candidates(text: str, article_url: str = "") -> list[str]:
    candidates: list[str] = []
    for url in _extract_urls(text):
        cleaned = _clean_url(url)
        if cleaned and cleaned not in candidates:
            candidates.append(cleaned)
    for prefix in ("来源：", "原文链接：", "原公告链接：", "详情见："):
        m = re.search(prefix + r"\s*([^\n。；]+)", text)
        if m:
            piece = _clean_url(m.group(1))
            if piece.startswith("http") and piece not in candidates:
                candidates.append(piece)
    if article_url and article_url not in candidates:
        candidates.append(article_url)
    return candidates


def _resolve_ddg_redirect(href: str) -> str:
    if href.startswith("//duckduckgo.com/l/?uddg="):
        parsed = urllib.parse.urlparse("https:" + href)
        query = urllib.parse.parse_qs(parsed.query)
        uddg = query.get("uddg", [""])[0]
        return urllib.parse.unquote(uddg)
    return href


def _search_public_source(title: str) -> tuple[str, str]:
    if not title.strip():
        return "", ""

    query = urllib.parse.quote(f'"{title.strip()}"')
    search_url = f"https://html.duckduckgo.com/html/?q={query}"
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        resp = requests.get(search_url, headers=headers, timeout=20)
        resp.raise_for_status()
    except Exception:
        return "", ""

    try:
        from bs4 import BeautifulSoup
    except Exception:
        return "", ""

    soup = BeautifulSoup(resp.text, "html.parser")
    title_chunks = _title_chunks(title)
    for node in soup.select(".result__a"):
        href = node.get("href") or ""
        href = _resolve_ddg_redirect(href)
        if not href.startswith("http"):
            continue
        text = _clean_text(node.get_text(" ", strip=True))
        if any(domain in href for domain in ["gongkaoleida.com", "weixin.qq.com", "zhihu.com", "bilibili.com", "animate.style", "wikipedia.org", "timeanddate.com", "calendar-365.com", "calendardate.com"]):
            continue
        if not any(chunk in text or chunk in href for chunk in title_chunks):
            continue
        if not any(domain in href for domain in [".gov.cn", ".edu.cn", ".org.cn", ".com.cn", ".cn/"]):
            continue
        return href, text
    return "", ""


def _utc_now_iso() -> str:
    return datetime.utcnow().isoformat()


def _should_skip_origin_url(url: str) -> bool:
    return any(domain in url for domain in ["gongkaoleida.com", "weixin.qq.com", "zhihu.com", "bilibili.com"])


async def _fetch_html_text(page: Page, url: str) -> tuple[str, str]:
    await _load_page(page, url)
    body_text = await _extract_best_text(
        page,
        [
            "article",
            "main",
            "div.article-content",
            "div.content",
            "div.detail-content",
            "div.article-detail",
            "div.mdn-content-box",
        ],
    )
    html = await page.content()
    return body_text, html


async def _extract_article_page_data(page: Page) -> dict[str, Any] | None:
    try:
        await page.wait_for_function(
            """() => {
                const el = document.querySelector('#article-app');
                const vm = el && el.__vue__;
                const pageData = vm && (vm.pageData || (vm.$data && vm.$data.pageData));
                return !!(pageData && (pageData.articleContent || pageData.articleInfo || pageData.examInfo));
            }""",
            timeout=12000,
        )
    except Exception:
        pass

    try:
        page_data = await page.evaluate(
            """() => {
                const el = document.querySelector('#article-app');
                const vm = el && el.__vue__;
                const pageData = vm && (vm.pageData || (vm.$data && vm.$data.pageData));
                if (!pageData) return null;
                return {
                    articleContent: pageData.articleContent || '',
                    articleInfo: pageData.articleInfo || {},
                    examInfo: pageData.examInfo || {},
                    examEnrollInfo: pageData.examEnrollInfo || {},
                    examTimeInfo: pageData.examTimeInfo || {},
                    recommendJobList: pageData.recommendJobList || [],
                    articleRelationList: pageData.articleRelationList || [],
                    isContentLimited: pageData.isContentLimited || false,
                    description: pageData.description || '',
                    keywords: pageData.keywords || '',
                    title: pageData.title || '',
                };
            }"""
        )
        return page_data if isinstance(page_data, dict) else None
    except Exception:
        return None


async def _extract_attachment_links(page: Page) -> list[dict[str, str]]:
    items = await page.locator("a[href]").evaluate_all(
        """
        (els) => els.map((a) => {
          const href = a.href || '';
          const text = (a.innerText || a.textContent || '').trim();
          return { href, text };
        })
        """
    )
    attachments: list[dict[str, str]] = []
    for item in items:
        href = str(item.get("href") or "").strip()
        text = _clean_text(item.get("text") or "")
        if not href:
            continue
        if _looks_like_attachment_url(href, text):
            attachments.append({"href": href, "text": text})
    deduped: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in attachments:
        href = item["href"]
        text = item.get("text", "")
        if "公考雷达APP" in text or "关于公考雷达" in text or "Sitemap" in text:
            continue
        if href in seen:
            continue
        seen.add(href)
        deduped.append(item)
    return deduped


class _ArticleContentParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.text_parts: list[str] = []
        self.attachments: list[dict[str, str]] = []
        self._current_anchor: dict[str, str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = {key: value or "" for key, value in attrs}
        if tag == "a":
            self._current_anchor = {
                "href": attrs_dict.get("href", ""),
                "oldsrc": attrs_dict.get("oldsrc", ""),
                "text": "",
            }
        if tag in {"p", "br", "div", "li"}:
            self.text_parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag == "a" and self._current_anchor is not None:
            text = _clean_text(self._current_anchor.get("text", ""))
            href = self._current_anchor.get("href", "")
            oldsrc = self._current_anchor.get("oldsrc", "")
            if oldsrc or _looks_like_attachment_url(href, text):
                self.attachments.append(
                    {
                        "href": href,
                        "oldsrc": oldsrc,
                        "text": text,
                    }
                )
            self._current_anchor = None

    def handle_data(self, data: str) -> None:
        if data:
            self.text_parts.append(data)
            if self._current_anchor is not None:
                self._current_anchor["text"] += data


def _parse_article_content_html(html: str) -> tuple[str, list[dict[str, str]]]:
    parser = _ArticleContentParser()
    parser.feed(html or "")
    text = _clean_text(" ".join(parser.text_parts))
    attachments: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for item in parser.attachments:
        key = (item.get("oldsrc", ""), item.get("text", ""))
        if key in seen:
            continue
        seen.add(key)
        attachments.append(item)
    return text, attachments


def _download_attachment(url: str) -> bytes:
    response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    response.raise_for_status()
    return response.content


def _parse_xlsx_attachment(content: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    rows: list[str] = []
    for sheet in wb.worksheets[:3]:
        rows.append(f"[sheet] {sheet.title}")
        for row in sheet.iter_rows(min_row=1, max_row=40, values_only=True):
            cells = [_clean_text(cell) for cell in row if _clean_text(cell)]
            if cells:
                rows.append(" | ".join(cells))
    return "\n".join(rows)


def _parse_docx_attachment(content: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for para in root.findall(".//w:p", ns):
        texts = [node.text or "" for node in para.findall(".//w:t", ns)]
        joined = _clean_text("".join(texts))
        if joined:
            paragraphs.append(joined)
    return "\n".join(paragraphs)


def _parse_text_attachment(content: bytes) -> str:
    for encoding in ("utf-8", "gb18030", "gbk"):
        try:
            return content.decode(encoding, errors="ignore")
        except Exception:
            continue
    return content.decode("utf-8", errors="ignore")


def _parse_attachment_content(url: str, content: bytes) -> str:
    lower_url = url.lower()
    try:
        if ".xlsx" in lower_url or ".xlsm" in lower_url:
            return _parse_xlsx_attachment(content)
        if ".docx" in lower_url:
            return _parse_docx_attachment(content)
        if any(ext in lower_url for ext in [".csv", ".txt", ".html", ".htm"]):
            return _parse_text_attachment(content)
    except Exception:
        return ""
    return ""


def _file_ext_from_attachment(href: str, oldsrc: str) -> str:
    candidate = (oldsrc or href or "").lower()
    match = re.search(r"\.([a-z0-9]{2,5})(?:$|[?#])", candidate)
    return match.group(1) if match else ""


def _collect_attachment_texts(attachments: list[dict[str, str]], *, max_count: int = 3) -> tuple[list[dict[str, str]], str]:
    parsed: list[dict[str, str]] = []
    text_blocks: list[str] = []
    for attachment in attachments[:max_count]:
        href = str(attachment.get("href") or "").strip()
        oldsrc = str(attachment.get("oldsrc") or "").strip()
        label = _clean_text(attachment.get("text") or "")
        if not href.startswith("http"):
            continue
        try:
            content = _download_attachment(href)
            parsed_text = _clean_text(_parse_attachment_content(href, content))
        except Exception:
            parsed_text = ""
        parsed.append(
            {
                "href": href,
                "oldsrc": oldsrc,
                "text": label,
                "parsed_text_preview": parsed_text[:1000],
            }
        )
        if parsed_text:
            text_blocks.append(f"[附件]{label or href}\n{parsed_text}")
    return parsed, "\n".join(text_blocks)


def _build_attachment_records(
    *,
    source_platform: str,
    source_id: str,
    scope: str,
    attachments: list[dict[str, str]],
    attachment_details: list[dict[str, str]],
) -> list[dict[str, object]]:
    detail_by_key: dict[tuple[str, str, str], dict[str, str]] = {}
    for item in attachment_details:
        key = (
            str(item.get("href") or ""),
            str(item.get("oldsrc") or ""),
            str(item.get("text") or ""),
        )
        detail_by_key[key] = item

    records: list[dict[str, object]] = []
    for item in attachments:
        href = str(item.get("href") or "")
        oldsrc = str(item.get("oldsrc") or "")
        text = _clean_text(item.get("text") or "")
        key = (href, oldsrc, text)
        detail = detail_by_key.get(key, {})
        parsed_text = str(detail.get("parsed_text_preview") or "")
        parse_status = "parsed" if parsed_text else ("hidden_token" if oldsrc else "metadata_only")
        record = {
            "event_source_platform": source_platform,
            "event_source_id": source_id,
            "attachment_scope": scope,
            "name": text,
            "href": href,
            "oldsrc": oldsrc,
            "file_ext": _file_ext_from_attachment(href, oldsrc),
            "parse_status": parse_status,
            "parsed_text": parsed_text,
            "raw_json": json.dumps(item, ensure_ascii=False),
        }
        records.append(record)
    return records


def _clean_summary_text(text: str) -> str:
    normalized = _clean_text(text)
    normalized = re.sub(r"(查看公告详情|关注|分享|来源：登录后展示来源|关注本场考试)", " ", normalized)
    normalized = _clean_text(normalized)
    return normalized


def _build_summary(article_text: str, body_text: str) -> str:
    primary = _clean_summary_text(article_text or body_text)
    if not primary:
        return ""
    chunks = re.split(r"[。！？]", primary)
    useful = [_clean_text(chunk) for chunk in chunks if _clean_text(chunk)]
    if useful:
        return "。".join(useful[:2])[:300]
    return primary[:300]


def _build_page_data_text(page_data: dict[str, Any]) -> str:
    parts: list[str] = []
    article_info = page_data.get("articleInfo") or {}
    exam_info = page_data.get("examInfo") or {}
    enroll_info = page_data.get("examEnrollInfo") or {}
    exam_time_info = page_data.get("examTimeInfo") or {}
    if isinstance(article_info, dict):
        for key in ["title", "origin", "author", "description"]:
            value = _clean_text(article_info.get(key) or "")
            if value:
                parts.append(f"{key}: {value}")
    if isinstance(exam_info, dict):
        for key in ["name", "title", "enroll_time_start", "enroll_time_end", "exam_time", "area_name", "category_name"]:
            value = _clean_text(exam_info.get(key) or "")
            if value:
                parts.append(f"{key}: {value}")
    if isinstance(enroll_info, dict):
        for key in ["start_time", "end_time", "enroll_time_start", "enroll_time_end", "status"]:
            value = _clean_text(enroll_info.get(key) or "")
            if value:
                parts.append(f"{key}: {value}")
    if isinstance(exam_time_info, dict):
        for key in ["written_time", "interview_time", "exam_time"]:
            value = _clean_text(exam_time_info.get(key) or "")
            if value:
                parts.append(f"{key}: {value}")

    jobs = page_data.get("recommendJobList") or []
    if isinstance(jobs, list) and jobs:
        job_names: list[str] = []
        for job in jobs[:50]:
            if not isinstance(job, dict):
                continue
            name = _clean_text(job.get("position") or job.get("title") or "")
            if name and name not in job_names:
                job_names.append(name)
        if job_names:
            parts.append("推荐职位: " + "、".join(job_names))
    return "\n".join(parts)


def _attachment_records_from_event(event: GongkaoEvent) -> list[dict[str, object]]:
    try:
        payload = json.loads(event.raw_json)
    except Exception:
        return []
    if not isinstance(payload, dict):
        return []

    records: list[dict[str, object]] = []
    article_attachments = payload.get("article_attachments") or []
    article_attachment_details = payload.get("article_attachment_details") or []
    if isinstance(article_attachments, list) and isinstance(article_attachment_details, list):
        records.extend(
            _build_attachment_records(
                source_platform=event.source_platform,
                source_id=event.source_id,
                scope="article",
                attachments=article_attachments,
                attachment_details=article_attachment_details,
            )
        )
    origin_attachments = payload.get("origin_attachments") or []
    origin_attachment_details = payload.get("origin_attachment_details") or []
    if isinstance(origin_attachments, list) and isinstance(origin_attachment_details, list):
        records.extend(
            _build_attachment_records(
                source_platform=event.source_platform,
                source_id=event.source_id,
                scope="origin",
                attachments=origin_attachments,
                attachment_details=origin_attachment_details,
            )
        )
    return records


async def _resolve_origin_source(page: Page, *, title: str, article_text: str, article_html: str) -> tuple[str, str]:
    candidates = _extract_origin_candidates(article_text)
    if not candidates:
        candidates = _extract_origin_candidates(article_html)

    for candidate in candidates:
        if not candidate.startswith("http") or _should_skip_origin_url(candidate):
            continue
        if not any(domain in candidate for domain in [".gov.cn", ".edu.cn", ".org.cn", ".com.cn", ".cn/"]):
            continue
        try:
            text, _ = await _fetch_html_text(page, candidate)
            if text and len(text) > 120 and _is_origin_text_match(title, text, candidate):
                return candidate, text
        except Exception:
            continue
    return "", ""


async def _resolve_origin_via_search(page: Page, title: str) -> tuple[str, str, str]:
    candidate_url, _ = _search_public_source(title)
    if not candidate_url:
        return "", "", ""
    try:
        text, html = await _fetch_html_text(page, candidate_url)
        if text and _is_origin_text_match(title, text, candidate_url):
            return candidate_url, text, html
    except Exception:
        return "", "", ""
    return "", "", ""


async def _enrich_event_origin(browser: Browser, event: GongkaoEvent, *, next_attempt: int) -> GongkaoEvent:
    page = await browser.new_page()
    try:
        origin_url = ""
        origin_text = ""
        origin_html = ""
        if event.article_url:
            try:
                article_text, article_html = await _fetch_html_text(page, event.article_url)
            except Exception:
                article_text, article_html = "", ""
            origin_url, origin_text = await _resolve_origin_source(
                page,
                title=event.title,
                article_text=article_text,
                article_html=article_html,
            )
            if not origin_url:
                origin_url, origin_text, origin_html = await _resolve_origin_via_search(page, event.title)
            elif origin_url:
                try:
                    await _load_page(page, origin_url)
                    origin_html = await page.content()
                except Exception:
                    origin_html = ""

        event.source_origin_url = origin_url
        event.source_origin_text = origin_text
        event.source_origin_html = origin_html
        event.origin_search_attempts = next_attempt
        event.origin_search_status = "found" if origin_url else ("not_found" if next_attempt >= 5 else "searched")
        event.origin_last_checked_at = _utc_now_iso()
        if origin_text:
            event.raw_text = "\n".join(part for part in [origin_text, event.raw_text] if part)
        return event
    finally:
        await page.close()


def _derive_status(card_status: str, registration_start: str, registration_deadline: str) -> str:
    today = datetime.now().date()
    if registration_deadline:
        try:
            deadline_date = datetime.strptime(registration_deadline, "%Y-%m-%d").date()
            if deadline_date < today:
                return "报名结束"
        except Exception:
            pass
    if card_status:
        return card_status
    if registration_start:
        try:
            start_date = datetime.strptime(registration_start, "%Y-%m-%d").date()
            if start_date > today:
                return "即将开始"
        except Exception:
            pass
    if registration_deadline:
        return "正在报名"
    return "报名结束"


async def _load_page(page: Page, url: str) -> None:
    await page.goto(url, wait_until="networkidle", timeout=60000)


async def _extract_best_text(page: Page, selectors: list[str]) -> str:
    for selector in selectors:
        locator = page.locator(selector).first
        try:
            if await locator.count() == 0:
                continue
            text = _clean_text(await locator.inner_text(timeout=5000))
            if text and len(text) > 60:
                return text
        except Exception:
            continue
    try:
        return _clean_text(await page.locator("body").inner_text(timeout=10000))
    except Exception:
        return ""


async def _parse_exam_list(page: Page, max_items: int) -> list[dict[str, str]]:
    items = await page.locator(".precise-notice-list ul.link-list > li").evaluate_all(
        """
        (els) => els.map((li) => {
          const text = (li.innerText || li.textContent || '').trim();
          const link = li.querySelector('a[href*="/exam/"]');
          const href = link ? link.href : '';
          const title = link ? (link.innerText || link.textContent || '').trim() : '';
          return { text, href, title };
        })
        """
    )
    return [item for item in items if item.get("href")][:max_items]


def _parse_card_meta(card_text: str) -> dict[str, str]:
    lines = [_clean_text(line) for line in card_text.splitlines() if _clean_text(line)]
    merged = " ".join(lines)
    title_match = re.search(r"\[([^\]]+)\]\s*\[([^\]]+)\]\s*(.+?)\s*/\s*(\d{4}-\d{2}-\d{2})", merged)
    if title_match:
        return {
            "region": title_match.group(1),
            "category": title_match.group(2),
            "title": title_match.group(3),
            "publish_date": title_match.group(4),
            "card_status": "正在报名" if "报名进行中" in merged else ("即将开始" if "报名未开始" in merged else ""),
        }
    return {
        "region": "",
        "category": "",
        "title": "",
        "publish_date": "",
        "card_status": "",
    }


async def _parse_event_detail(browser: Browser, list_item: dict[str, str], *, use_origin_search: bool) -> GongkaoEvent:
    card_text = list_item.get("text", "")
    meta = _parse_card_meta(card_text)
    exam_url = list_item["href"]
    source_id = exam_url.rstrip("/").split("/")[-1]

    page = await browser.new_page()
    await _load_page(page, exam_url)
    body_text = await _extract_best_text(
        page,
        [
            "div.article-introduce",
            "div.mdn-content-box",
            "div.article-detail",
        ],
    )
    article_link = await page.locator("a.title-related").get_attribute("href")
    article_url = f"{BASE_URL}{article_link}" if article_link and article_link.startswith("/") else (article_link or "")
    summary = _build_summary("", body_text)

    org_name = meta["title"]
    if "公告" in org_name:
        org_name = org_name.split("公告")[0].strip()
    org_name = re.sub(r"(招聘|招录|公开招聘|面向社会公开招聘).*$", "", org_name).strip(" ：:，,")

    job_count_match = re.search(r"共招(\d+)人", body_text)
    job_count = int(job_count_match.group(1)) if job_count_match else None

    publish_time = ""
    publish_match = re.search(r"(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}", body_text)
    if publish_match:
        publish_time = f"{publish_match.group(1)}T00:00:00"
    elif meta["publish_date"]:
        publish_time = f"{meta['publish_date']}T00:00:00"

    article_text = ""
    origin_url = ""
    origin_text = ""
    origin_html = ""
    article_attachments: list[dict[str, str]] = []
    origin_attachments: list[dict[str, str]] = []
    article_attachment_details: list[dict[str, str]] = []
    origin_attachment_details: list[dict[str, str]] = []
    attachment_text = ""
    page_data = None
    page_data_text = ""
    if article_url:
        article_text, article_html = await _fetch_html_text(page, article_url)
        article_attachments = await _extract_attachment_links(page)
        page_data = await _extract_article_page_data(page)

        if page_data and page_data.get("articleContent"):
            article_content_text, article_content_attachments = _parse_article_content_html(page_data["articleContent"])
            article_text = "\n".join(part for part in [article_text, article_content_text] if part)
            article_attachments = article_attachments + article_content_attachments
        if page_data:
            page_data_text = _build_page_data_text(page_data)
        summary = _build_summary(article_text, body_text)
        article_attachment_details, article_attachment_text = _collect_attachment_texts(article_attachments)
        if article_attachment_text:
            attachment_text = article_attachment_text
        if use_origin_search:
            origin_url, origin_text = await _resolve_origin_source(
                page,
                title=meta["title"] or list_item.get("title") or "",
                article_text=article_text,
                article_html=article_html,
            )
            if not origin_url:
                origin_url, origin_text, origin_html = await _resolve_origin_via_search(
                    page,
                    meta["title"] or list_item.get("title") or "",
                )
            if origin_url:
                try:
                    await _load_page(page, origin_url)
                    origin_html = await page.content()
                    origin_attachments = await _extract_attachment_links(page)
                    origin_attachment_details, origin_attachment_text = _collect_attachment_texts(origin_attachments)
                    if origin_attachment_text:
                        attachment_text = "\n".join(part for part in [attachment_text, origin_attachment_text] if part)
                except Exception:
                    origin_attachments = []

    raw_text = "\n".join(part for part in [origin_text, article_text, page_data_text, attachment_text, body_text] if part)
    registration_start, registration_deadline = _extract_registration_dates(raw_text, publish_time)
    if not registration_start and page_data:
        exam_info = page_data.get("examInfo") or {}
        registration_start = _normalize_date(str(exam_info.get("enroll_time_start") or ""), None)
    exam_date = _extract_exam_date(raw_text, publish_time)
    qualification = _extract_qualification(raw_text)
    major_requirements = _extract_major_requirements(raw_text)
    status = _derive_status(meta["card_status"], registration_start, registration_deadline)
    origin_search_status = "found" if origin_url else ("searched" if use_origin_search else "pending")
    origin_search_attempts = 1 if use_origin_search else 0
    origin_last_checked_at = _utc_now_iso() if use_origin_search else ""

    hash_basis = "|".join(["gongkaoleida", source_id])
    hash_id = hashlib.md5(hash_basis.encode("utf-8")).hexdigest()

    raw_json = json.dumps(
        {
            "list_item": list_item,
            "meta": meta,
            "article_url": article_url,
            "origin_url": origin_url,
            "origin_text": origin_text,
            "origin_html": origin_html,
            "body_text": body_text,
            "article_text": article_text,
            "article_attachments": article_attachments,
            "origin_attachments": origin_attachments,
            "article_attachment_details": article_attachment_details,
            "origin_attachment_details": origin_attachment_details,
            "page_data": page_data,
        },
        ensure_ascii=False,
    )
    await page.close()
    return GongkaoEvent(
        source_platform="gongkaoleida",
        source_id=source_id,
        title=meta["title"] or list_item.get("title") or source_id,
        region=meta["region"],
        category=meta["category"],
        org_name=org_name,
        job_count=job_count,
        qualification=qualification,
        major_requirements=major_requirements,
        registration_start=registration_start,
        registration_deadline=registration_deadline,
        exam_date=exam_date,
        status=status,
        publish_time=publish_time,
        source_url=exam_url,
        article_url=article_url,
        summary=summary,
        raw_text=raw_text,
        raw_json=raw_json,
        hash_id=hash_id,
        source_origin_url=origin_url,
        source_origin_text=origin_text,
        source_origin_html=origin_html,
        origin_search_status=origin_search_status,
        origin_search_attempts=origin_search_attempts,
        origin_last_checked_at=origin_last_checked_at,
    )


async def crawl_gongkaoleida(*, max_items: int, page_no: int, list_path: str, use_origin_search: bool) -> list[GongkaoEvent]:
    async with async_playwright() as p:
        browser = await p.chromium.launch(channel="chrome", headless=True)
        page = await browser.new_page(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
        )
        normalized_path = list_path if list_path.startswith("/") else f"/{list_path}"
        list_url = f"{BASE_URL}{normalized_path}?page={page_no}"
        await _load_page(page, list_url)
        items = await _parse_exam_list(page, max_items=max_items)
        await page.close()
        events: list[GongkaoEvent] = []
        for item in items:
            try:
                event = await _parse_event_detail(browser, item, use_origin_search=use_origin_search)
                events.append(event)
            except Exception as exc:
                print(f"[warn] failed to parse {item.get('href')}: {exc}")
        await browser.close()
        return events


def _should_attempt_origin_search(store: ContentStore, *, source_platform: str, source_id: str, enabled: bool) -> tuple[bool, int]:
    if not enabled:
        return False, 0
    state = store.get_gongkao_event_search_state(source_platform=source_platform, source_id=source_id)
    if not state:
        return True, 1
    current_attempts = int(state.get("origin_search_attempts") or 0)
    status = str(state.get("origin_search_status") or "")
    source_origin_url = str(state.get("source_origin_url") or "")
    if source_origin_url or status == "found":
        return False, current_attempts
    if current_attempts >= 5 or status == "not_found":
        return False, current_attempts
    return True, current_attempts + 1


def main() -> None:
    parser = argparse.ArgumentParser(description="Crawl gongkaoleida exam announcements into gongkao_events.")
    parser.add_argument("--db", default=str(CONFIG.database_path), help="SQLite database path.")
    parser.add_argument("--max_items", type=int, default=10, help="Maximum exam cards to crawl from the page.")
    parser.add_argument("--page", type=int, default=1, help="List page number to crawl.")
    parser.add_argument(
        "--category",
        default="all",
        help="Category alias: all/gwy/sydw/guoqi/teacher/medical/xuandiao.",
    )
    parser.add_argument(
        "--list_path",
        default="",
        help="Optional custom gongkaoleida list path, e.g. /exam_search/1-78 . Overrides --category.",
    )
    parser.add_argument(
        "--use_origin_search",
        action="store_true",
        help="Try to resolve the original source announcement from the title/article text.",
    )
    args = parser.parse_args()

    list_path = args.list_path or CATEGORY_PATHS.get(args.category.lower(), "/exam_search")
    store = ContentStore(Path(args.db))
    raw_events = asyncio.run(
        crawl_gongkaoleida(
            max_items=args.max_items,
            page_no=args.page,
            list_path=list_path,
            use_origin_search=False,
        )
    )

    events: list[GongkaoEvent] = []
    if args.use_origin_search:
        async def enrich_events() -> list[GongkaoEvent]:
            enriched: list[GongkaoEvent] = []
            async with async_playwright() as p:
                browser = await p.chromium.launch(channel="chrome", headless=True)
                for event in raw_events:
                    should_search, next_attempt = _should_attempt_origin_search(
                        store,
                        source_platform=event.source_platform,
                        source_id=event.source_id,
                        enabled=True,
                    )
                    if not should_search:
                        state = store.get_gongkao_event_search_state(
                            source_platform=event.source_platform,
                            source_id=event.source_id,
                        ) or {}
                        event.source_origin_url = str(state.get("source_origin_url") or "")
                        event.source_origin_text = str(state.get("source_origin_text") or "")
                        event.source_origin_html = str(state.get("source_origin_html") or "")
                        event.origin_search_status = str(state.get("origin_search_status") or event.origin_search_status)
                        event.origin_search_attempts = next_attempt
                        event.origin_last_checked_at = str(state.get("origin_last_checked_at") or event.origin_last_checked_at)
                        enriched.append(event)
                        continue
                    refreshed = await _enrich_event_origin(browser, event, next_attempt=next_attempt)
                    enriched.append(refreshed)
                await browser.close()
            return enriched

        events = asyncio.run(enrich_events())
    else:
        events = raw_events

    for event in events:
        store.upsert_gongkao_event(event.__dict__)
        store.replace_gongkao_event_attachments(
            source_platform=event.source_platform,
            source_id=event.source_id,
            attachments=_attachment_records_from_event(event),
        )
    print(f"Saved {len(events)} gongkao events into {args.db} from {list_path}?page={args.page}")
    for event in events[:5]:
        print(
            f"- [{event.status}] {event.region} / {event.category} / {event.title} / "
            f"deadline={event.registration_deadline or 'N/A'} / jobs={event.job_count or 'N/A'}"
        )


if __name__ == "__main__":
    main()
